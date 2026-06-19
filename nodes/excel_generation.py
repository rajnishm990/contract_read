import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import (
    FIELD_DISPLAY_NAMES,
    HIGH_CONFIDENCE,
    LOW_CONFIDENCE,
    OUTPUT_DIR,
)
from models.state import ContractState


_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_BLUE_HDR = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_LIGHT_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_WHITE_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")


def _conf_fill(conf: float) -> PatternFill:
    if conf >= HIGH_CONFIDENCE:
        return _GREEN
    if conf >= LOW_CONFIDENCE:
        return _YELLOW
    return _RED


def _autosize(ws, max_width: int = 80) -> None:
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, max_width)


def _header_row(ws, headers: List[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _BLUE_HDR
        c.font = _WHITE_FONT
        c.alignment = _CENTER


#  Sheet 1: Extracted Fields 

def _sheet_fields(wb: openpyxl.Workbook, fields: Dict) -> None:
    ws = wb.active
    ws.title = "Extracted Fields"

    _header_row(ws, ["Field", "Extracted Value", "Confidence", "Page Ref"])
    row = 2

    for key, display in FIELD_DISPLAY_NAMES.items():
        if key == "price_details":
            continue  # handled on Sheet 2

        fd = fields.get(key, {})
        value = fd.get("value")
        conf = float(fd.get("confidence", 0.0))
        page_ref = fd.get("page_ref", "")

        ws.cell(row=row, column=1, value=display).font = _BOLD
        vc = ws.cell(row=row, column=2, value=str(value) if value is not None else "NOT FOUND")
        vc.alignment = _WRAP
        vc.fill = _conf_fill(conf) if value is not None else _RED
        ws.cell(row=row, column=3, value=f"{conf:.0%}").fill = _conf_fill(conf)
        ws.cell(row=row, column=4, value=str(page_ref) if page_ref else "—")
        row += 1

    _autosize(ws)


# Sheet 2: Rate Card 

def _sheet_rate_card(wb: openpyxl.Workbook, fields: Dict) -> None:
    ws = wb.create_sheet("Rate Card")
    fd = fields.get("price_details", {})
    conf = float(fd.get("confidence", 0.0))
    page_ref = fd.get("page_ref", "")
    items: Any = fd.get("value", [])

    ws.cell(row=1, column=1, value="Confidence").font = _BOLD
    ws.cell(row=1, column=2, value=f"{conf:.0%}").fill = _conf_fill(conf)
    ws.cell(row=1, column=3, value="Page Reference").font = _BOLD
    ws.cell(row=1, column=4, value=str(page_ref) if page_ref else "—")

    if not isinstance(items, list) or not items:
        ws.cell(row=3, column=1, value="No rate card data extracted.").font = Font(italic=True)
        return

    # Dynamic headers from union of all item keys
    all_keys: List[str] = []
    seen_keys: set = set()
    for item in items:
        for k in item.keys():
            if k not in seen_keys:
                all_keys.append(k)
                seen_keys.add(k)

    _header_row(ws, [k.replace("_", " ").title() for k in all_keys], row=3)
    for r, item in enumerate(items, 4):
        for col, k in enumerate(all_keys, 1):
            ws.cell(row=r, column=col, value=str(item.get(k, ""))).alignment = _WRAP

    _autosize(ws)


# Sheet 3: Raw OCR Text 

def _sheet_raw_text(wb: openpyxl.Workbook, raw_text_by_page: Dict) -> None:
    ws = wb.create_sheet("Raw OCR Text")
    _header_row(ws, ["Page", "Extracted Text"])

    for row, (pg_num, text) in enumerate(sorted(raw_text_by_page.items()), 2):
        ws.cell(row=row, column=1, value=f"Page {pg_num + 1}").font = _BOLD
        ws.cell(row=row, column=2, value=text).alignment = _WRAP

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 120


#  Node 

def excel_generation_node(state: ContractState) -> dict:
    log = list(state.get("processing_log", []))
    fields = state.get("extracted_fields", {})
    raw_text = state.get("raw_text_by_page", {})

    wb = openpyxl.Workbook()
    _sheet_fields(wb, fields)
    _sheet_rate_card(wb, fields)
    _sheet_raw_text(wb, raw_text)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = Path(state.get("original_filename", "contract")).stem
    filename = f"{stem}_extracted_{ts}.xlsx"
    out_path = OUTPUT_DIR / filename
    wb.save(str(out_path))

    log.append(f"Excel saved: {filename}")
    return {
        **state,
        "excel_output_path": str(out_path),
        "processing_log": log,
        "current_step": "done",
    }
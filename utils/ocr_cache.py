from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl

from config.settings import OUTPUT_DIR

_OCR_SHEET = "Raw OCR Text"


def find_cached_ocr(original_filename: str) -> Optional[Path]:
    """Return the most-recent Excel in OUTPUT_DIR that:
    - has a stem starting with the uploaded file's stem
    - contains a 'Raw OCR Text' sheet
    Returns None if no match is found.
    """
    stem = Path(original_filename).stem.lower()
    candidates = sorted(OUTPUT_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for xlsx in candidates:
        xlsx_stem = xlsx.stem.lower()
        # Extract the original filename portion — everything before "_extracted_"
        if "_extracted_" in xlsx_stem:
            original_part = xlsx_stem.split("_extracted_")[0]
        else:
            original_part = xlsx_stem
        if original_part != stem:
            continue
        try:
            wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
            if _OCR_SHEET in wb.sheetnames:
                wb.close()
                return xlsx
            wb.close()
        except Exception:
            continue
    return None


def load_ocr_from_excel(xlsx_path: Path) -> Tuple[Dict[int, str], str]:
    """Read the 'Raw OCR Text' sheet and return (raw_text_by_page, full_text)."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[_OCR_SHEET]

    raw_text_by_page: Dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        page_label = str(row[0])  # e.g. "Page 1"
        text = str(row[1]) if row[1] is not None else ""
        try:
            page_num = int(page_label.replace("Page", "").strip()) - 1
        except ValueError:
            continue
        raw_text_by_page[page_num] = text

    wb.close()

    full_text = "\n\n".join(
        f"--- Page {pg + 1} ---\n{txt}"
        for pg, txt in sorted(raw_text_by_page.items())
    )
    return raw_text_by_page, full_text